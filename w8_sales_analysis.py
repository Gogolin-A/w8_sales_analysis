import collections
import pandas
import openpyxl


all_browsers = collections.defaultdict(int)
all_products = collections.defaultdict(int)
visit_date = collections.defaultdict(int)
sales_date = collections.defaultdict(int)
all_products_male = collections.defaultdict(int)
all_products_female = collections.defaultdict(int)

data_log = pandas.read_excel('logs.xlsx', sheet_name='log').to_dict(orient='records')
#считаем браузеры и товары
for elem in data_log:
    all_browsers[elem['Браузер']] += 1
    for item in elem['Купленные товары'].split(','):
        all_products[item] += 1
        if elem['Пол'] == 'м':
            all_products_male[item] += 1
        if elem['Пол'] == 'ж':
            all_products_female[item] += 1

popular_browsers = collections.Counter(all_browsers).most_common(7)
popular_products = collections.Counter(all_products).most_common(7)
#Заполняем популярные браузеры и товары
wb = openpyxl.load_workbook(filename='report.xlsx')
sheet = wb['Лист1']
for i in range(5, 12):
    sheet.cell(row=i, column=1).value = popular_browsers[i-5][0]
for i in range(5, 12):
    sheet.cell(row=i, column=2).value = popular_browsers[i-5][1]
for i in range(19, 26):
    sheet.cell(row=i, column=1).value = popular_products[i-19][0]
for i in range(19, 26):
    sheet.cell(row=i, column=2).value = popular_products[i-19][1]
wb.save(filename='report.xlsx')
#Заполняем количество посещений по месяцам
row_cnt = 5
for browser in popular_browsers:
    for elem in data_log:
        if browser[0] == elem['Браузер']:
            visit_date[elem['Дата посещения'].month] += 1
    wb = openpyxl.load_workbook(filename='report.xlsx')
    sheet = wb['Лист1']
    for j in range(3, 15):
        sheet.cell(row=row_cnt, column=j).value = collections.Counter(visit_date)[j-2]
    wb.save(filename='report.xlsx')
    visit_date = collections.defaultdict(int)
    row_cnt += 1
#Заполняем количество продаж по месяцам
row_cnt = 19
for product in popular_products:
    for elem in data_log:
        for item in elem['Купленные товары'].split(','):
            if product[0] == item:
                sales_date[elem['Дата посещения'].month] += 1
    wb = openpyxl.load_workbook(filename='report.xlsx')
    sheet = wb['Лист1']
    for j in range(3, 15):
        sheet.cell(row=row_cnt, column=j).value = collections.Counter(sales_date)[j-2]
    wb.save(filename='report.xlsx')
    sales_date = collections.defaultdict(int)
    row_cnt += 1
#заполняем самые популярные и непопулярные товары
wb = openpyxl.load_workbook(filename='report.xlsx')
sheet = wb['Лист1']
sheet.cell(row=31, column=2).value = collections.Counter(all_products_male).most_common(1)[0][0]
sheet.cell(row=32, column=2).value = collections.Counter(all_products_female).most_common(1)[0][0]
sheet.cell(row=33, column=2).value = collections.Counter(all_products_male).most_common()[::-1][0][0]
sheet.cell(row=34, column=2).value = collections.Counter(all_products_female).most_common()[::-1][0][0]
wb.save(filename='report.xlsx')
